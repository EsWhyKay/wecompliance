#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate static HTML pages from Excel sources.

Inputs (defaults):
  - weclac.xlsx   : compliance-style sheets (Food/Supp/Drug/Feed + links)
  - compare.xlsx  : comparison-style sheets (dynamic columns, no links)

Outputs:
  - One HTML page per worksheet across both workbooks.
  - index.html Home with two sections:
      1) WecLac菌株全球合规性 (cards to compliance pages)
      2) 其他菌株对比 (cards to comparison pages)

Key behaviors:
  - Region grouping: 亚洲 / 美洲 / 欧洲 / 大洋洲 / 其他
    Priority: Region column > region_map.csv > inference from country text/flag.
    USP always grouped as 其他.
  - In status cells, standalone tokens 'Y'/'N' are replaced with inline SVG icons.
  - Comparison pages: column titles come from header row; no links row.
  - Default collapsed.
"""

import argparse
import csv
import html
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from urllib.parse import urlparse

import openpyxl

REGION_ORDER = ["亚洲", "美洲", "欧洲", "大洋洲", "其他"]

# Explicit country-to-region overrides (business rule)
# Turkey and Egypt are treated as Europe; USP is Other.
COUNTRY_REGION_MAP = {
    # Asia
    '中国': '亚洲', '泰国': '亚洲', '印尼': '亚洲', '印度尼西亚': '亚洲', '马来西亚': '亚洲',
    '新加坡': '亚洲', '菲律宾': '亚洲', '越南': '亚洲', '柬埔寨': '亚洲', '伊朗': '亚洲',
    '日本': '亚洲', '韩国': '亚洲',
    # Americas
    '美国': '美洲', '加拿大': '美洲', '巴西': '美洲', '智利': '美洲',
    # Europe (per requested grouping)
    '欧盟': '欧洲', 'EU': '欧洲', 'European Union': '欧洲',
    '英国': '欧洲', 'UK': '欧洲', 'United Kingdom': '欧洲',
    '俄罗斯': '欧洲', 'Russia': '欧洲', 'Russian Federation': '欧洲',
    '土耳其': '欧洲', 'Turkey': '欧洲',
    '埃及': '欧洲', 'Egypt': '欧洲',
    # Oceania
    '澳洲': '大洋洲', '澳大利亚': '大洋洲', 'Australia': '大洋洲',
    '新西兰': '大洋洲', 'New Zealand': '大洋洲',
    # Other
    '南非': '其他', 'South Africa': '其他', 'USP': '其他',
}

def normalize_country_name(name: str) -> str:
    """Normalize country label by stripping flag emojis and extra punctuation/whitespace."""
    import re
    s = (name or '').strip()
    # Remove flag emojis (regional indicator symbols) and other emoji
    s = re.sub(r'[🇦-🇿]{2}', '', s)
    s = re.sub(r'[𐀀-􏿿]', '', s)
    # Remove leading bullets/arrows and surrounding whitespace
    s = s.strip().lstrip('•·-–—→↗')
    return s.strip()

COMPLIANCE_REQUIRED = [
    "Country",
    "食品", "f_link",
    "保健品/膳食补充剂", "s_link",
    "药品", "d_link",
    "动物食品", "fe_link",
]

DEFAULT_COLUMNS = {
    "country": "Country",
    "food": "食品",
    "f_link": "f_link",
    "supp": "保健品/膳食补充剂",
    "s_link": "s_link",
    "drug": "药品",
    "d_link": "d_link",
    "feed": "动物食品",
    "fe_link": "fe_link",
    "region": "Region",  # optional
}

TEAM = [
    {
        "name": "Xiaowen Wang",
        "email": "xiaowen.wang@wecare-bio.com",
        "role": "Consultant, Eurasian Plate",
        "photo": "assets/photos/xiaowen.wang.svg",
        "links": [],
    },
    {
        "name": "Yukun Sun",
        "email": "kay.sun@wecare-life.com",
        "role": "Manager, Pacific Plate",
        "photo": "assets/photos/kay.sun.svg",
        "links": [
            {"label": "LinkedIn", "href": "http://www.linkedin.com/in/yu-kun-sun"}
        ],
    },
    {
        "name": "Yixuan Fan",
        "email": "yixuan.fan@wecare-bio.com",
        "role": "Consultant, American Plate",
        "photo": "assets/photos/yixuan.fan.svg",
        "links": [],
    },
]


HOME_TAGLINES = {
    'BC99': '<i>Bacillus coagulans</i> / <i>Weizmannia coagulans</i> / <i>Heyndrickxia coagulans</i>',
    'Akk11': '<i>Akkermansia muciniphila</i>',
    'BLa80': '<i>Bifidobacterium animalis</i> subsp. <i>lactis</i>',
    'LRa05': '<i>Lacticaseibacillus rhamnosus</i>',
}




def slugify(name: str) -> str:
    import re
    s = (name or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "sheet"


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def escape(s: Optional[str]) -> str:
    if s is None:
        return ""
    return html.escape(str(s), quote=False)


def load_region_map(csv_path: Path) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    if not csv_path.exists():
        return mapping
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            c = (row.get("country") or "").strip()
            r = (row.get("region") or "").strip()
            if c and r:
                mapping[c] = r
    return mapping


def normalize_region(r: str) -> str:
    r = (r or "").strip()
    if r in ("亚洲", "美洲", "欧洲", "大洋洲"):
        return r
    rl = r.lower()
    if rl in ("asia",):
        return "亚洲"
    if rl in ("americas", "america", "north america", "south america", "latin america"):
        return "美洲"
    if rl in ("europe", "eu"):
        return "欧洲"
    if rl in ("oceania",):
        return "大洋洲"
    return "其他"


def flag_to_iso2(flag: str) -> str:
    # Convert regional indicator symbols to ISO2.
    # Returns "" if not a valid flag.
    try:
        if len(flag) < 2:
            return ""
        codepoints = [ord(ch) for ch in flag[:2]]
        # Regional indicator A starts at 0x1F1E6
        if all(0x1F1E6 <= cp <= 0x1F1FF for cp in codepoints):
            return chr(codepoints[0] - 0x1F1E6 + ord('A')) + chr(codepoints[1] - 0x1F1E6 + ord('A'))
        return ""
    except Exception:
        return ""


def infer_region(country_raw: str) -> str:
    # USP forced other
    if not country_raw:
        return "其他"
    s = str(country_raw).strip()
    if "USP" in s.upper():
        return "其他"

    # Strip leading flag emoji if present
    iso2 = ""
    if s and len(s) >= 2:
        iso2 = flag_to_iso2(s[:2])

    # Minimal ISO2 -> region map for common entries
    iso_region = {
        # Asia
        "CN": "亚洲", "JP": "亚洲", "KR": "亚洲", "TW": "亚洲", "HK": "亚洲", "MO": "亚洲",
        "SG": "亚洲", "MY": "亚洲", "TH": "亚洲", "VN": "亚洲", "ID": "亚洲", "IN": "亚洲",
        "AE": "亚洲", "SA": "亚洲", "IL": "亚洲",
        # Americas
        "US": "美洲", "CA": "美洲", "MX": "美洲", "BR": "美洲", "AR": "美洲", "CL": "美洲", "PE": "美洲", "CO": "美洲",
        # Europe
        "GB": "欧洲", "FR": "欧洲", "DE": "欧洲", "IT": "欧洲", "ES": "欧洲", "NL": "欧洲", "BE": "欧洲", "SE": "欧洲", "NO": "欧洲", "CH": "欧洲",
        "EU": "欧洲",
        # Oceania
        "AU": "大洋洲", "NZ": "大洋洲",
    }
    if iso2 in iso_region:
        return iso_region[iso2]

    # Text heuristics (Chinese + English names / common abbreviations)
    t = s
    # Remove leading flag for text checks
    if iso2:
        t = t[2:].strip()

    t_upper = t.upper()
    t_lower = t.lower()

    # Europe
    if any(k in t for k in ["欧盟", "欧洲", "英国", "法国", "德国", "意大利", "西班牙", "荷兰", "瑞士", "瑞典", "挪威", "比利时", "爱尔兰", "奥地利", "丹麦", "芬兰", "葡萄牙", "波兰", "捷克", "匈牙利", "罗马尼亚", "希腊"]):
        return "欧洲"
    if any(k in t_upper for k in ["EU", "UK", "GB"]):
        return "欧洲"
    if any(k in t_lower for k in [
        "european union", "europe", "united kingdom", "england", "scotland", "wales",
        "france", "germany", "italy", "spain", "netherlands", "switzerland", "sweden",
        "norway", "belgium", "ireland", "austria", "denmark", "finland", "portugal",
        "poland", "czech", "hungary", "romania", "greece"
    ]):
        return "欧洲"

    # Americas
    if any(k in t for k in ["美国", "加拿大", "墨西哥", "巴西", "阿根廷", "智利", "秘鲁", "哥伦比亚"]):
        return "美洲"
    if any(k in t_lower for k in [
        "united states", "u.s.", "usa", "america", "canada", "mexico", "brazil", "argentina",
        "chile", "peru", "colombia"
    ]):
        return "美洲"

    # Oceania
    if any(k in t for k in ["澳大利亚", "新西兰"]):
        return "大洋洲"
    if any(k in t_lower for k in ["australia", "new zealand", "nz"]):
        return "大洋洲"

    # Asia
    if any(k in t for k in ["中国", "日本", "韩国", "台湾", "香港", "澳门", "新加坡", "马来西亚", "泰国", "越南", "印度", "印尼", "阿联酋", "沙特", "以色列"]):
        return "亚洲"
    if any(k in t_lower for k in [
        "china", "japan", "korea", "south korea", "taiwan", "hong kong", "macau", "macao",
        "singapore", "malaysia", "thailand", "vietnam", "india", "indonesia", "uae",
        "united arab emirates", "saudi", "israel"
    ]):
        return "亚洲"

    return "其他"


def classify_region(country: str, region_cell: Optional[str], region_map: Dict[str, str]) -> str:
    # 1) Country explicit overrides (after normalization)
    c0 = normalize_country_name(str(country or ''))
    if c0 in COUNTRY_REGION_MAP:
        return normalize_region(COUNTRY_REGION_MAP[c0])
    for k, v in COUNTRY_REGION_MAP.items():
        if k and k in c0:
            return normalize_region(v)
    # 2) Explicit region cell (if present)
    if region_cell and str(region_cell).strip():
        r0 = str(region_cell).strip()
        # accept common English labels
        r_map = {
            'Asia': '亚洲', 'Asian': '亚洲',
            'Americas': '美洲', 'America': '美洲', 'North America': '美洲', 'South America': '美洲',
            'Europe': '欧洲', 'European': '欧洲',
            'Oceania': '大洋洲',
        }
        r0 = r_map.get(r0, r0)
        return normalize_region(r0)
    # 3) region_map.csv lookup (normalized country)
    if c0 in region_map:
        return normalize_region(region_map[c0])
    if str(country or '').strip() in region_map:
        return normalize_region(region_map[str(country).strip()])
    return '其他'


def svg_to_inline(svg_text: str) -> str:
    s = (svg_text or "").strip()
    if s.startswith("<?xml"):
        s = s.split("?>", 1)[-1].lstrip()
    return s


def build_external_icon_svg() -> str:
    return (
        "<span class=\"ext\">"
        "<svg viewBox=\"0 0 24 24\" fill=\"none\" xmlns=\"http://www.w3.org/2000/svg\">"
        "<path d=\"M14 3h7v7\" stroke=\"currentColor\" stroke-width=\"2\" stroke-linecap=\"round\" stroke-linejoin=\"round\"/>"
        "<path d=\"M10 14L21 3\" stroke=\"currentColor\" stroke-width=\"2\" stroke-linecap=\"round\" stroke-linejoin=\"round\"/>"
        "<path d=\"M21 14v7H3V3h7\" stroke=\"currentColor\" stroke-width=\"2\" stroke-linecap=\"round\" stroke-linejoin=\"round\"/>"
        "</svg>"
        "</span>"
    )


def render_status(cell: Optional[str], svg_yes: str, svg_no: str) -> str:
    """Replace standalone Y/N tokens with inline SVG. Preserve text and line breaks."""
    if cell is None:
        return "-"
    raw = str(cell)
    s = raw.replace("\r\n", "\n").replace("\r", "\n")
    if not s.strip():
        return "-"

    import re

    # Match a standalone Y/N token (including full-width Ｙ/Ｎ) not glued to letters/numbers.
    # Followed by whitespace OR end-of-string.
    pat = re.compile(r"(?<![A-Za-z0-9])([YN\uFF39\uFF2E])(?=\s|$)", re.IGNORECASE)

    parts: List[str] = []
    last = 0
    for m in pat.finditer(s):
        if m.start() > last:
            parts.append(escape(s[last:m.start()]))
        ch = m.group(1)
        ch_u = ch.upper()
        icon = svg_yes if ch_u in ("Y", "\uFF39") else svg_no
        parts.append(f"<span class=\"status\"><span class=\"ico\">{icon}</span></span>")
        last = m.end()
    if last < len(s):
        parts.append(escape(s[last:]))

    out = "".join(parts)
    out = out.replace("\n", "<br>")
    return out if out.strip() else "-"


def render_link(url: Optional[str], svg_external: str) -> str:
    u = ("" if url is None else str(url)).strip()
    if not u:
        return "-"

    # Allow "url | title" input, but do not fetch web titles.
    safe_u = html.escape(u, quote=True)
    title = ""
    raw_url = u
    if "|" in u:
        parts = [p.strip() for p in u.split("|", 1)]
        raw_url = parts[0]
        safe_u = html.escape(parts[0], quote=True)
        title = parts[1] if len(parts) > 1 else ""

    def derive_host(raw: str) -> str:
        try:
            p = urlparse(raw)
            if p.netloc:
                return p.netloc
            p2 = urlparse("https://" + raw)
            if p2.netloc:
                return p2.netloc
            raw2 = raw.split("//", 1)[-1]
            return raw2.split("/", 1)[0]
        except Exception:
            raw2 = (raw or "").split("//", 1)[-1]
            return raw2.split("/", 1)[0]

    label_txt = title.strip() if title.strip() else derive_host(raw_url)
    label = f"{svg_external} {escape(label_txt)}".strip()
    return f"<a href=\"{safe_u}\" target=\"_blank\" rel=\"noopener\">{label}</a>"


def html_shell(doc_title: str,
               h1_title: str,
               subtitle: str,
               nav_html: str,
               body_html: str,
               team_html: str,
               show_controls: bool) -> str:
    tpl = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>__DOC_TITLE__</title>
  <style>
    :root {
      --bg-main: #f4f1ec;
      --bg-card: #fbfaf7;
      --bg-expand: #ffffff;
      --border-soft: #ddd6cc;
      --text-main: #2f2f2f;
      --text-muted: #746f67;
      --accent: #7c6c55;
      --link: #5f5342;
      --chip: #efe9df;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      background: var(--bg-main);
      color: var(--text-main);
    }
    header { text-align: center; padding: 28px 16px 10px; }
    header h1 { margin: 0; font-size: 36px; letter-spacing: 1px; }
    header h2 { margin: 10px 0 0; font-size: 15px; font-weight: 400; color: var(--text-muted); }

    .topbar { max-width: 1180px; margin: 0 auto; padding: 10px 16px 0; }
    .nav {
      display: flex; gap: 8px; flex-wrap: wrap;
      justify-content: center;
      margin: 8px 0 14px;
    }
    .nav a {
      display: inline-block;
      padding: 8px 12px;
      border: 1px solid var(--border-soft);
      background: var(--bg-card);
      border-radius: 999px;
      text-decoration: none;
      color: var(--accent);
      font-size: 13px;
    }
    .nav a.active {
      background: var(--chip);
      font-weight: 600;
    }

    .controls {
      max-width: 1180px;
      margin: 0 auto 14px;
      padding: 0 16px;
      display: flex;
      gap: 12px;
      flex-wrap: wrap;
    }
    .controls input {
      flex: 1;
      min-width: 240px;
      padding: 10px 12px;
      border: 1px solid var(--border-soft);
      border-radius: 10px;
      background: #fff;
      font-size: 14px;
      outline: none;
    }
    .controls button {
      padding: 10px 14px;
      border: 1px solid var(--border-soft);
      background: var(--bg-card);
      border-radius: 10px;
      cursor: pointer;
      color: var(--accent);
      font-size: 13px;
    }

    .container { max-width: 1180px; margin: 0 auto 44px; padding: 0 16px; }

    .section-title {
      margin: 28px 0 16px;
      padding: 4px 0;
      font-weight: 800;
      color: var(--accent);
      font-size: 16px;
      line-height: 1.6;
    }

    .sheetcards {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 12px;
    }
    .sheetcard {
      display: flex;
      gap: 12px;
      align-items: center;
      padding: 14px 16px;
      background: var(--bg-card);
      border: 1px solid var(--border-soft);
      border-radius: 14px;
      text-decoration: none;
      color: var(--text-main);
    }
    .sheetcard .img {
      width: 44px;
      height: 44px;
      border-radius: 12px;
      border: 1px solid var(--border-soft);
      background: #fff;
      flex: 0 0 auto;
    }
    .sheetcard .meta { min-width: 0; }
    .sheetcard .t { font-weight: 800; color: var(--accent); font-size: 16px; }
    .sheetcard .d { margin-top: 6px; color: var(--text-muted); font-size: 13px; }
    .sheetcard:hover { box-shadow: 0 1px 0 rgba(0,0,0,0.05); }

    details.group {
      background: var(--bg-card);
      border: 1px solid var(--border-soft);
      border-radius: 14px;
      margin-bottom: 12px;
      overflow: hidden;
    }
    details.group > summary {
      padding: 14px 16px;
      cursor: pointer;
      list-style: none;
      font-weight: 700;
      color: var(--accent);
      display: flex;
      align-items: center;
      justify-content: space-between;
    }
    details.group > summary::-webkit-details-marker { display: none; }
    .badge { font-size: 12px; color: var(--text-muted); background: var(--chip); padding: 4px 10px; border-radius: 999px; border: 1px solid var(--border-soft); }

    details.country {
      background: transparent;
      border-top: 1px solid var(--border-soft);
    }
    details.country > summary {
      padding: 12px 16px;
      font-size: 15px;
      cursor: pointer;
      list-style: none;
      display: flex;
      align-items: center;
      justify-content: space-between;
      font-weight: 600;
      color: var(--text-main);
    }
    details.country > summary::-webkit-details-marker { display: none; }

    .content {
      background: var(--bg-expand);
      border-top: 1px solid var(--border-soft);
      padding: 14px 16px 16px;
    }

    /* Force-hide details content when collapsed to prevent blank reserved space */
    details.country:not([open]) > .content { display: none !important; }

    .grid {
      display: grid;
      grid-template-columns: repeat(4, 1fr);
      gap: 12px;
      font-size: 13.5px;
    }
    .cell-title { font-weight: 700; color: var(--accent); }
    .cell { line-height: 1.55; white-space: pre-wrap; word-break: break-word; }

    .links {
      margin-top: 10px;
      padding-top: 10px;
      border-top: 1px dashed var(--border-soft);
      display: grid;
      grid-template-columns: repeat(4, 1fr);
      gap: 12px;
      font-size: 12.5px;
      color: var(--text-muted);
      align-items: start;
    }

    a { color: var(--link); text-decoration: none; }
    a:hover { text-decoration: underline; }

    .status { display: inline-flex; gap: 8px; align-items: flex-start; }
    .status .ico { width: 18px; height: 18px; display: inline-flex; flex: 0 0 auto; margin-top: 1px; }
    .status .ico svg { width: 18px; height: 18px; display: block; }

    .ext { display: inline-flex; align-items: center; gap: 6px; }
    .ext svg { width: 14px; height: 14px; }

    footer {
      max-width: 1180px;
      margin: 0 auto 64px;
      padding: 0 16px;
      color: var(--text-muted);
    }
    .team {
      background: var(--bg-card);
      border: 1px solid var(--border-soft);
      border-radius: 14px;
      padding: 16px;
    }
    .team h3 { margin: 0 0 12px; color: var(--accent); font-size: 16px; }
    .members {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 12px;
    }
    @media (max-width: 900px) {
      .grid, .links { grid-template-columns: 1fr !important; }
      .members { grid-template-columns: 1fr; }
      .sheetcards { grid-template-columns: 1fr; }
    }

    .member {
      background: #fff;
      border: 1px solid var(--border-soft);
      border-radius: 12px;
      padding: 12px;
      display: flex;
      gap: 12px;
      align-items: flex-start;
    }
    .avatar {
      width: 56px;
      height: 56px;
      border-radius: 12px;
      border: 1px solid var(--border-soft);
      overflow: hidden;
      background: var(--bg-main);
      flex: 0 0 auto;
      display: grid;
      place-items: center;
    }
    .avatar img { width: 100%; height: 100%; object-fit: cover; display: block; }
    .mmeta { font-size: 13px; }
    .mmeta .name { font-weight: 700; color: var(--text-main); }
    .mmeta .roleline { color: var(--accent); font-size: 12px; margin-top: 4px; line-height: 1.5; }
    .mmeta .roleline a { color: var(--link); }
    .mmeta .sep { color: var(--border-soft); padding: 0 6px; }

    .hint { text-align: center; color: var(--text-muted); font-size: 12px; margin-top: 8px; }
  </style>
</head>
<body>
  <header>
    <h1>__H1__</h1>
    <h2>__SUBTITLE__</h2>
  </header>

  <div class="topbar">
    <div class="nav">__NAV__</div>
  </div>

  __CONTROLS__

  <div class="container" id="container">
    __BODY__
  </div>

  <footer>
    __TEAM__
    <div class="hint">If you have any questions, please contact us.</div>
  </footer>

  <script>
    function toggleAll(open) {
      const countries = document.querySelectorAll('details.country');
      const groups = document.querySelectorAll('details.group');
      if (open) {
        groups.forEach(g => { g.open = true; g.setAttribute('open',''); });
        countries.forEach(c => { c.open = true; c.setAttribute('open',''); });
      } else {
        // Close children first to avoid layout glitches in some browsers
        countries.forEach(c => { c.open = false; c.removeAttribute('open'); });
        groups.forEach(g => { g.open = false; g.removeAttribute('open'); });
      }
    }

    function filterList() {
      const q = (document.getElementById('search')?.value || '').toLowerCase().trim();
      const groups = document.querySelectorAll('details.group');

      groups.forEach(g => {
        let anyVisible = false;
        const countries = g.querySelectorAll('details.country');
        countries.forEach(c => {
          const hay = (c.getAttribute('data-hay') || '').toLowerCase();
          const show = (!q) || hay.includes(q);
          c.style.display = show ? '' : 'none';
          if (show) anyVisible = true;
        });
        g.style.display = anyVisible ? '' : 'none';
        if (q && anyVisible) g.open = true;
      });
    }

    // Ensure region expansion does not auto-expand all countries after a global collapse.
    document.addEventListener('click', function(e){
      const sum = e.target.closest('details.group > summary');
      if(!sum) return;
      const g = sum.parentElement;
      // When a region is opened by user click, collapse all its country items by default.
      setTimeout(()=>{
        if(g && g.open){
          g.querySelectorAll('details.country').forEach(c=>{c.open=false; c.removeAttribute('open');});
        }
      }, 0);
    });
  </script>
</body>
</html>
"""

    controls_html = ""
    if show_controls:
        controls_html = (
            "<div class=\"controls\">"
            "<input id=\"search\" placeholder=\"Search country or content…\" oninput=\"filterList()\" />"
            "<button onclick=\"toggleAll(true)\">Expand all</button>"
            "<button onclick=\"toggleAll(false)\">Collapse all</button>"
            "</div>"
        )

    return (tpl
            .replace("__DOC_TITLE__", escape(doc_title))
            .replace("__H1__", escape(h1_title))
            .replace("__SUBTITLE__", escape(subtitle))
            .replace("__NAV__", nav_html)
            .replace("__CONTROLS__", controls_html)
            .replace("__BODY__", body_html)
            .replace("__TEAM__", team_html)
            )


def build_team_html(placeholder_svg: str) -> str:
    parts = ["<div class=\"team\">", "<h3>Team</h3>", "<div class=\"members\">"]
    for m in TEAM:
        email_txt = escape(m["email"])
        line2_parts = [f"<a href=\"mailto:{email_txt}\">{email_txt}</a>"]
        for l in m.get("links", []):
            line2_parts.append(
                f"<a href=\"{html.escape(l['href'], quote=True)}\" target=\"_blank\" rel=\"noopener\">{escape(l['label'])}</a>"
            )
        line2 = " <span class=\"sep\">·</span> ".join(line2_parts)

        photo = html.escape(m["photo"], quote=True)
        img = (
            f"<img src=\"{photo}\" alt=\"{escape(m['name'])}\" "
            f"onerror=\"this.onerror=null;this.src='{placeholder_svg}';\" />"
        )

        parts.append(
            "<div class=\"member\">"
            f"<div class=\"avatar\">{img}</div>"
            "<div class=\"mmeta\">"
            f"<div class=\"name\">{escape(m['name'])}</div>"
            f"<div class=\"roleline\">{escape(m['role'])}<br>{line2}</div>"
            "</div>"
            "</div>"
        )
    parts.append("</div></div>")
    return "".join(parts)


def build_nav_html(pages: List[Tuple[str, str]], active_href: str) -> str:
    # UX requirement: only keep the Home button in the navbar.
    cls = "active" if active_href == "index.html" else ""
    return f"<a class=\"{cls}\" href=\"index.html\">Home</a>"


def detect_sheet_kind(headers: List[str]) -> str:
    headers_set = set([h for h in headers if h is not None])
    if all(h in headers_set for h in COMPLIANCE_REQUIRED):
        return "compliance"
    if DEFAULT_COLUMNS["country"] in headers_set:
        return "compare"
    return "unknown"


def sheet_rows_compliance(ws, headers: List[str]) -> List[Dict[str, str]]:
    idx = {h: i for i, h in enumerate(headers)}

    def get(row, col):
        if col not in idx:
            return None
        return row[idx[col]]

    rows_out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() for v in r):
            continue
        country = get(r, DEFAULT_COLUMNS["country"])
        if not country or not str(country).strip():
            continue
        rows_out.append({
            "country": str(country).strip(),
            "region": get(r, DEFAULT_COLUMNS["region"]),
            "food": get(r, DEFAULT_COLUMNS["food"]),
            "f_link": get(r, DEFAULT_COLUMNS["f_link"]),
            "supp": get(r, DEFAULT_COLUMNS["supp"]),
            "s_link": get(r, DEFAULT_COLUMNS["s_link"]),
            "drug": get(r, DEFAULT_COLUMNS["drug"]),
            "d_link": get(r, DEFAULT_COLUMNS["d_link"]),
            "feed": get(r, DEFAULT_COLUMNS["feed"]),
            "fe_link": get(r, DEFAULT_COLUMNS["fe_link"]),
        })
    return rows_out


def sheet_rows_dynamic(ws, headers: List[str]) -> Tuple[List[Dict[str, str]], List[str]]:
    idx = {h: i for i, h in enumerate(headers)}
    if DEFAULT_COLUMNS["country"] not in idx:
        raise SystemExit(f"Sheet '{ws.title}' missing column: {DEFAULT_COLUMNS['country']}")

    display_cols = [h for h in headers if h not in (DEFAULT_COLUMNS["country"], DEFAULT_COLUMNS["region"]) and h is not None]

    rows_out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() for v in r):
            continue
        country = r[idx[DEFAULT_COLUMNS["country"]]]
        if not country or not str(country).strip():
            continue

        obj: Dict[str, str] = {
            "country": str(country).strip(),
            "region": r[idx[DEFAULT_COLUMNS["region"]]] if DEFAULT_COLUMNS["region"] in idx else None,
        }
        for c in display_cols:
            obj[c] = r[idx[c]] if c in idx else None
        rows_out.append(obj)

    return rows_out, display_cols


def build_body_compliance(rows: List[Dict[str, str]], region_map: Dict[str, str], svg_yes: str, svg_no: str) -> str:
    ext_icon = build_external_icon_svg()

    grouped: Dict[str, List[Dict[str, str]]] = {k: [] for k in REGION_ORDER}
    for r in rows:
        region = classify_region(r["country"], r.get("region"), region_map)
        grouped.setdefault(region, []).append(r)

    for reg in grouped:
        grouped[reg].sort(key=lambda x: x["country"])

    parts: List[str] = []
    for reg in REGION_ORDER:
        items = grouped.get(reg, [])
        if not items:
            continue
        parts.append(
            f"<details class=\"group\"><summary><span>{escape(reg)}</span><span class=\"badge\">{len(items)} countries</span></summary>"
        )
        for r in items:
            food_html = render_status(r.get("food"), svg_yes, svg_no)
            supp_html = render_status(r.get("supp"), svg_yes, svg_no)
            drug_html = render_status(r.get("drug"), svg_yes, svg_no)
            feed_html = render_status(r.get("feed"), svg_yes, svg_no)

            f_link_html = render_link(r.get("f_link"), ext_icon)
            s_link_html = render_link(r.get("s_link"), ext_icon)
            d_link_html = render_link(r.get("d_link"), ext_icon)
            fe_link_html = render_link(r.get("fe_link"), ext_icon)

            hay = " | ".join([
                r.get("country", "") or "",
                str(r.get("food") or ""),
                str(r.get("supp") or ""),
                str(r.get("drug") or ""),
                str(r.get("feed") or ""),
                str(r.get("f_link") or ""),
                str(r.get("s_link") or ""),
                str(r.get("d_link") or ""),
                str(r.get("fe_link") or ""),
            ])
            hay_attr = html.escape(hay, quote=True)

            parts.append(
                f"<details class=\"country\" data-hay=\"{hay_attr}\">"
                f"<summary><span>{escape(r['country'])}</span></summary>"
                "<div class=\"content\">"
                "<div class=\"grid\">"
                "<div class=\"cell-title\">食品</div>"
                "<div class=\"cell-title\">保健品/膳食补充剂</div>"
                "<div class=\"cell-title\">药品</div>"
                "<div class=\"cell-title\">动物食品</div>"
                f"<div class=\"cell\">{food_html}</div>"
                f"<div class=\"cell\">{supp_html}</div>"
                f"<div class=\"cell\">{drug_html}</div>"
                f"<div class=\"cell\">{feed_html}</div>"
                "</div>"
                "<div class=\"links\">"
                f"<div>{f_link_html}</div>"
                f"<div>{s_link_html}</div>"
                f"<div>{d_link_html}</div>"
                f"<div>{fe_link_html}</div>"
                "</div>"
                "</div>"
                "</details>"
            )
        parts.append("</details>")

    return "".join(parts)


def build_body_dynamic(rows: List[Dict[str, str]],
                       display_cols: List[str],
                       region_map: Dict[str, str],
                       svg_yes: str,
                       svg_no: str) -> str:
    grouped: Dict[str, List[Dict[str, str]]] = {k: [] for k in REGION_ORDER}
    for r in rows:
        region = classify_region(r["country"], r.get("region"), region_map)
        grouped.setdefault(region, []).append(r)

    for reg in grouped:
        grouped[reg].sort(key=lambda x: x["country"])

    col_n = max(1, len(display_cols))

    parts: List[str] = []
    for reg in REGION_ORDER:
        items = grouped.get(reg, [])
        if not items:
            continue
        parts.append(
            f"<details class=\"group\"><summary><span>{escape(reg)}</span><span class=\"badge\">{len(items)} countries</span></summary>"
        )
        for r in items:
            titles_html = "".join([f"<div class=\"cell-title\">{escape(c)}</div>" for c in display_cols])
            cells_html = "".join([f"<div class=\"cell\">{render_status(r.get(c), svg_yes, svg_no)}</div>" for c in display_cols])

            hay = " | ".join([r.get("country", "")] + [str(r.get(c) or "") for c in display_cols])
            hay_attr = html.escape(hay, quote=True)

            parts.append(
                f"<details class=\"country\" data-hay=\"{hay_attr}\">"
                f"<summary><span>{escape(r['country'])}</span></summary>"
                "<div class=\"content\">"
                f"<div class=\"grid\" style=\"grid-template-columns: repeat({col_n}, 1fr);\">"
                f"{titles_html}{cells_html}"
                "</div>"
                "</div>"
                "</details>"
            )
        parts.append("</details>")

    return "".join(parts)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--weclac", default="weclac.xlsx", help="Compliance workbook")
    ap.add_argument("--compare", default="compare.xlsx", help="Comparison workbook")
    ap.add_argument("--out", default="build", help="Output directory")
    args = ap.parse_args()

    here = Path(__file__).resolve().parent
    out_dir = (here / args.out).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    svg_yes = svg_to_inline(read_text(here / "assets/icons/yes.svg"))
    svg_no = svg_to_inline(read_text(here / "assets/icons/no.svg"))

    placeholder_svg = (
        "data:image/svg+xml;utf8,"
        "<svg xmlns='http://www.w3.org/2000/svg' width='120' height='120'>"
        "<rect width='100%25' height='100%25' rx='18' ry='18' fill='%23f4f1ec'/>"
        "<circle cx='60' cy='48' r='18' fill='%23ddd6cc'/>"
        "<rect x='26' y='74' width='68' height='30' rx='15' fill='%23ddd6cc'/>"
        "</svg>"
    )

    region_map = load_region_map(here / "region_map.csv")

    wb_a = openpyxl.load_workbook(here / args.weclac)
    wb_b = openpyxl.load_workbook(here / args.compare)

    pages: List[Tuple[str, str]] = [("Home", "index.html")]

    # Collect sheets and create unique labels/hrefs
    # Compliance pages: href = <sheet>.html
    # Compare pages: href = compare-<sheet>.html (avoid conflicts)

    compliance_sheets: List[str] = []
    compare_sheets: List[str] = []

    # Compliance
    for sn in wb_a.sheetnames:
        ws = wb_a[sn]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if detect_sheet_kind(headers) == "compliance":
            compliance_sheets.append(sn)

    # Compare
    for sn in wb_b.sheetnames:
        ws = wb_b[sn]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if detect_sheet_kind(headers) == "compare":
            compare_sheets.append(sn)

    # Build pages list with duplicate label handling
    used_labels = set(["Home"])
    for sn in compliance_sheets:
        label = sn
        href = f"{sn}.html"
        pages.append((label, href))
        used_labels.add(label)

    for sn in compare_sheets:
        label = sn if sn not in used_labels else f"{sn} (Compare)"
        href = f"compare-{sn}.html"
        pages.append((label, href))
        used_labels.add(label)

    # Build compliance pages
    for sn in compliance_sheets:
        ws = wb_a[sn]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = sheet_rows_compliance(ws, headers)
        body = build_body_compliance(rows, region_map, svg_yes, svg_no)
        nav = build_nav_html(pages, f"{sn}.html")
        team_html = build_team_html(placeholder_svg)
        html_text = html_shell(
            doc_title=f"WeCompliance - {sn}",
            h1_title=sn,
            subtitle="WeCompliance · WecLac菌株全球合规性",
            nav_html=nav,
            body_html=body,
            team_html=team_html,
            show_controls=True,
        )
        (out_dir / f"{sn}.html").write_text(html_text, encoding="utf-8")

    # Build compare pages
    for sn in compare_sheets:
        ws = wb_b[sn]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows, display_cols = sheet_rows_dynamic(ws, headers)
        body = build_body_dynamic(rows, display_cols, region_map, svg_yes, svg_no)
        nav = build_nav_html(pages, f"compare-{sn}.html")
        team_html = build_team_html(placeholder_svg)
        html_text = html_shell(
            doc_title=f"WeCompliance - {sn}",
            h1_title=sn,
            subtitle="WeCompliance · 国际菌株对比",
            nav_html=nav,
            body_html=body,
            team_html=team_html,
            show_controls=True,
        )
        (out_dir / f"compare-{sn}.html").write_text(html_text, encoding="utf-8")

    # Build index.html
    nav = build_nav_html(pages, "index.html")

    def make_card(label: str, href: str) -> str:
        icon_slug = slugify(label)
        icon_path = f"assets/sheets/{icon_slug}.svg"
        tagline = HOME_TAGLINES.get(label, "Open page")
        return (
            f"<a class=\"sheetcard\" href=\"{html.escape(href, quote=True)}\">"
            f"<img class=\"img\" src=\"{html.escape(icon_path, quote=True)}\" alt=\"\" onerror=\"this.onerror=null;this.src='assets/sheets/default.svg';\">"
            "<div class=\"meta\">"
            f"<div class=\"t\">{escape(label)}</div>"
            f"<div class=\"d\">{tagline}</div>"
            "</div></a>"
        )

    # Section 1: WecLac compliance cards
    sec1 = ["<div class=\"section-title\">WecLac菌株全球合规性</div>", "<div class=\"sheetcards\">"]
    for sn in compliance_sheets:
        sec1.append(make_card(sn, f"{sn}.html"))
    sec1.append("</div>")

    # Section 2: Compare cards
    sec2 = ["<div class=\"section-title\">国际菌株对比</div>", "<div class=\"sheetcards\">"]
    for sn in compare_sheets:
        # Always display sheet name only
        sec2.append(make_card(sn, f"compare-{sn}.html"))
    sec2.append("</div>")

    home_body = "".join(sec1 + sec2)

    team_html = build_team_html(placeholder_svg)
    index_html = html_shell(
        doc_title="WeCompliance",
        h1_title="WeCompliance",
        subtitle="Probiotics Compliance Advice Around the World",
        nav_html=nav,
        body_html=home_body,
        team_html=team_html,
        show_controls=False,
    )
    (out_dir / "index.html").write_text(index_html, encoding="utf-8")

    # Ensure default collapse all: details are closed by default; no action needed.

    print("Generated to:", out_dir)


if __name__ == "__main__":
    main()

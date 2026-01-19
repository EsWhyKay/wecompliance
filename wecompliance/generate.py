#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate static HTML pages from data.xlsx.

- One page per worksheet (e.g., BC99.html)
- index.html as a landing page with links to each worksheet page
- Region grouping: 亚洲 / 美洲 / 欧洲 / 大洋洲 (fallback: 其他)
- Cells (食品/保健品/药品/动物食品) may start with 'Y' or 'N' and will be rendered with inline SVG icons.

Usage:
  python generate.py
  python generate.py --excel data.xlsx --out .
"""

import argparse
import csv
import html
import os
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from urllib.parse import urlparse

import openpyxl

REGION_ORDER = ["亚洲", "美洲", "欧洲", "大洋洲", "其他"]

DEFAULT_COLUMNS = {
    "country": "Country",
    "food": "食品",
    "f_link": "f_link",
    "supp": "保健品/膳食补充剂",
    "s_link": "s_link",
    "drug": "药品",
    "d_link": "d_link",
    "feed": "动物食品",
    "fe_link": "fe_link",
    # optional
    "region": "Region",
}

TEAM = [
    {
        "name": "Xiaowen Wang",
        "email": "xiaowen.wang@wecare-bio.com",
        "role": "Consultant, Eurasian Plate",
        "note": "",
        "photo": "assets/photos/xiaowen.wang.svg",
        "links": [],
    },
    {
        "name": "Yukun Sun",
        "email": "kay.sun@wecare-life.com",
        "role": "Manager, Pacific Plate",
        "note": "",
        "photo": "assets/photos/kay.sun.svg",
        "links": [
            {"label": "LinkedIn", "href": "http://www.linkedin.com/in/yu-kun-sun"}
        ],
    },
    {
        "name": "Yixuan Fan",
        "email": "yixuan.fan@wecare-bio.com",
        "role": "Consultant, American Plate",
        "note": "",
        "photo": "assets/photos/yixuan.fan.svg",
        "links": [],
    },
]


def slugify(name: str) -> str:
    """Filesystem-safe slug for sheet icon filenames."""
    import re
    s = (name or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "sheet"


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def load_region_map(csv_path: Path) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    if not csv_path.exists():
        return mapping
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            c = (row.get("country") or "").strip()
            r = (row.get("region") or "").strip()
            if c and r:
                mapping[c] = r
    return mapping


def escape(s: Optional[str]) -> str:
    if s is None:
        return ""
    return html.escape(str(s), quote=False)


def normalize_region(r: str) -> str:
    r = (r or "").strip()
    if r in ("亚洲", "美洲", "欧洲", "大洋洲"):
        return r
    return "其他" if r else "其他"


def classify_region(country: str, region_cell: Optional[str], region_map: Dict[str, str]) -> str:
    if region_cell and str(region_cell).strip():
        return normalize_region(str(region_cell).strip())
    if country in region_map:
        return normalize_region(region_map[country])
    return "其他"


def render_status(cell: Optional[str], svg_yes: str, svg_no: str) -> str:
    """Render Y/N markers inside a cell as inline SVG icons.

    User rule:
      - In the 4 status columns, any "Y" or "N" that is followed by whitespace
        (space/tab/etc.) should be replaced with the corresponding SVG.
      - This should work across line breaks (i.e., multiple lines in one cell).

    Implementation:
      - Operate line-by-line (preserve line breaks).
      - Replace every standalone marker character (Y/N) that is immediately
        followed by whitespace. The rest of the text remains unchanged.
    """
    if cell is None:
        return "-"

    raw = str(cell)
    s = raw.replace("\r\n", "\n").replace("\r", "\n")
    if not s.strip():
        return "-"

    import re

    pat = re.compile(r"([YN])(?=\s)")

    def _replace_in_line(line: str) -> str:
        if not line:
            return ""
        parts: List[str] = []
        last = 0
        for m in pat.finditer(line):
            # Escape text before match
            if m.start() > last:
                parts.append(escape(line[last:m.start()]))
            ch = m.group(1)
            icon = svg_yes if ch == "Y" else svg_no
            parts.append(f"<span class=\"status\"><span class=\"ico\">{icon}</span></span>")
            last = m.end()
        if last < len(line):
            parts.append(escape(line[last:]))
        return "".join(parts)

    out_lines: List[str] = []
    for line in s.split("\n"):
        out_lines.append(_replace_in_line(line))

    # Join lines preserving breaks
    html_block = "<br>".join(out_lines).strip()
    return html_block if html_block else "-"


def render_link(url: Optional[str], svg_external: str) -> str:
    u = ("" if url is None else str(url)).strip()
    if not u:
        return "-"
    safe_u = html.escape(u, quote=True)
    # Title is NOT fetched from the web (offline). Users may optionally provide it as: "url | title".
    title = ""
    if "|" in u:
        parts = [p.strip() for p in u.split("|", 1)]
        if parts[0]:
            safe_u = html.escape(parts[0], quote=True)
        title = parts[1] if len(parts) > 1 else ""

    def _derive_title(raw_url: str) -> str:
        """Offline link label rule:

        Show only the part up to the first '/' (i.e., the host).
        Examples:
          https://example.com/a/b -> example.com
          example.com/a -> example.com
        """
        try:
            u0 = (raw_url or "").strip()
            if not u0:
                return ""
            p = urlparse(u0)
            host = (p.netloc or "").strip()
            if host:
                return host
            # Handle URLs without scheme
            p2 = urlparse("https://" + u0)
            host2 = (p2.netloc or "").strip()
            if host2:
                return host2
            # Fallback: naive split
            u0 = u0.split("//", 1)[-1]
            return u0.split("/", 1)[0]
        except Exception:
            u0 = (raw_url or "").strip()
            u0 = u0.split("//", 1)[-1]
            return u0.split("/", 1)[0]

    title_txt = escape(title) if title else escape(_derive_title(u.split("|", 1)[0].strip()))
    label = f"{svg_external} {title_txt}".strip()
    return f"<a href=\"{safe_u}\" target=\"_blank\" rel=\"noopener\">{label}</a>"


def html_shell(page_title: str, subtitle: str, nav_html: str, body_html: str, team_html: str) -> str:
    # Use __PLACEHOLDERS__ to avoid curly brace escaping pain.
    tpl = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>__PAGE_TITLE__</title>
  <style>
    :root {
      --bg-main: #f4f1ec;
      --bg-card: #fbfaf7;
      --bg-expand: #ffffff;
      --border-soft: #ddd6cc;
      --text-main: #2f2f2f;
      --text-muted: #746f67;
      --accent: #7c6c55;
      --link: #5f5342;
      --chip: #efe9df;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      background: var(--bg-main);
      color: var(--text-main);
    }
    header { text-align: center; padding: 28px 16px 10px; }
    header h1 { margin: 0; font-size: 36px; letter-spacing: 1px; }
    header h2 { margin: 10px 0 0; font-size: 15px; font-weight: 400; color: var(--text-muted); }

    .topbar { max-width: 1180px; margin: 0 auto; padding: 10px 16px 0; }
    .nav {
      display: flex; gap: 8px; flex-wrap: wrap;
      justify-content: center;
      margin: 8px 0 14px;
    }
    .nav a {
      display: inline-block;
      padding: 8px 12px;
      border: 1px solid var(--border-soft);
      background: var(--bg-card);
      border-radius: 999px;
      text-decoration: none;
      color: var(--accent);
      font-size: 13px;
    }
    .nav a.active {
      background: var(--chip);
      font-weight: 600;
    }

    .controls {
      max-width: 1180px;
      margin: 0 auto 14px;
      padding: 0 16px;
      display: flex;
      gap: 12px;
      flex-wrap: wrap;
    }
    .controls input {
      flex: 1;
      min-width: 240px;
      padding: 10px 12px;
      border: 1px solid var(--border-soft);
      border-radius: 10px;
      background: #fff;
      font-size: 14px;
      outline: none;
    }
    .controls button {
      padding: 10px 14px;
      border: 1px solid var(--border-soft);
      background: var(--bg-card);
      border-radius: 10px;
      cursor: pointer;
      color: var(--accent);
      font-size: 13px;
    }

    .container { max-width: 1180px; margin: 0 auto 44px; padding: 0 16px; }

    .sheetcards {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 12px;
    }
    .sheetcard {
      display: flex;
      gap: 12px;
      align-items: center;
      padding: 14px 16px;
      background: var(--bg-card);
      border: 1px solid var(--border-soft);
      border-radius: 14px;
      text-decoration: none;
      color: var(--text-main);
    }
    .sheetcard .img {
      width: 44px;
      height: 44px;
      border-radius: 12px;
      border: 1px solid var(--border-soft);
      background: #fff;
      flex: 0 0 auto;
    }
    .sheetcard .meta { min-width: 0; }
    .sheetcard .t {
      font-weight: 800;
      color: var(--accent);
      font-size: 16px;
    }
    .sheetcard .d {
      margin-top: 6px;
      color: var(--text-muted);
      font-size: 13px;
    }
    .sheetcard:hover { box-shadow: 0 1px 0 rgba(0,0,0,0.05); }

    details.group {
      background: var(--bg-card);
      border: 1px solid var(--border-soft);
      border-radius: 14px;
      margin-bottom: 12px;
      overflow: hidden;
    }
    details.group > summary {
      padding: 14px 16px;
      cursor: pointer;
      list-style: none;
      font-weight: 700;
      color: var(--accent);
      display: flex;
      align-items: center;
      justify-content: space-between;
    }
    details.group > summary::-webkit-details-marker { display: none; }
    .badge { font-size: 12px; color: var(--text-muted); background: var(--chip); padding: 4px 10px; border-radius: 999px; border: 1px solid var(--border-soft); }

    details.country {
      background: transparent;
      border-top: 1px solid var(--border-soft);
    }
    details.country > summary {
      padding: 12px 16px;
      font-size: 15px;
      cursor: pointer;
      list-style: none;
      display: flex;
      align-items: center;
      justify-content: space-between;
      font-weight: 600;
      color: var(--text-main);
    }
    details.country > summary::-webkit-details-marker { display: none; }

    .content {
      background: var(--bg-expand);
      border-top: 1px solid var(--border-soft);
      padding: 14px 16px 16px;
    }

    .grid {
      display: grid;
      grid-template-columns: repeat(4, 1fr);
      gap: 12px;
      font-size: 13.5px;
    }
    .cell-title { font-weight: 700; color: var(--accent); }
    .cell { line-height: 1.55; white-space: pre-wrap; word-break: break-word; }

    .links {
      margin-top: 10px;
      padding-top: 10px;
      border-top: 1px dashed var(--border-soft);
      display: grid;
      grid-template-columns: repeat(4, 1fr);
      gap: 12px;
      font-size: 12.5px;
      color: var(--text-muted);
      align-items: start;
    }

    a { color: var(--link); text-decoration: none; }
    a:hover { text-decoration: underline; }

    .status { display: inline-flex; gap: 8px; align-items: flex-start; }
    .status .ico { width: 18px; height: 18px; display: inline-flex; flex: 0 0 auto; margin-top: 1px; }
    .status .ico svg { width: 18px; height: 18px; display: block; }
    .status .stxt { display: inline-block; }

    .ext { display: inline-flex; align-items: center; gap: 6px; }
    .ext svg { width: 14px; height: 14px; }

    footer {
      max-width: 1180px;
      margin: 0 auto 64px;
      padding: 0 16px;
      color: var(--text-muted);
    }
    .team {
      background: var(--bg-card);
      border: 1px solid var(--border-soft);
      border-radius: 14px;
      padding: 16px;
    }
    .team h3 { margin: 0 0 12px; color: var(--accent); font-size: 16px; }
    .members {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 12px;
    }
    @media (max-width: 900px) {
      .grid, .links { grid-template-columns: 1fr; }
      .members { grid-template-columns: 1fr; }
      .sheetcards { grid-template-columns: 1fr; }
    }

    .member {
      background: #fff;
      border: 1px solid var(--border-soft);
      border-radius: 12px;
      padding: 12px;
      display: flex;
      gap: 12px;
      align-items: flex-start;
    }
    .avatar {
      width: 56px;
      height: 56px;
      border-radius: 12px;
      border: 1px solid var(--border-soft);
      overflow: hidden;
      background: var(--bg-main);
      flex: 0 0 auto;
      display: grid;
      place-items: center;
    }
    .avatar img { width: 100%; height: 100%; object-fit: cover; display: block; }
    .mmeta { font-size: 13px; }
    .mmeta .name { font-weight: 700; color: var(--text-main); }
    .mmeta .roleline { color: var(--accent); font-size: 12px; margin-top: 4px; line-height: 1.5; }
    .mmeta .roleline a { color: var(--link); }
    .mmeta .sep { color: var(--border-soft); padding: 0 6px; }

    .hint { text-align: center; color: var(--text-muted); font-size: 12px; margin-top: 8px; }
  </style>
</head>
<body>
  <header>
    <h1>__H1__</h1>
    <h2>__SUBTITLE__</h2>
  </header>

  <div class="topbar">
    <div class="nav">__NAV__</div>
  </div>

  <div class="controls">
    <input id="search" placeholder="Search country or content…" oninput="filterList()" />
    <button onclick="toggleAll(true)">Expand all</button>
    <button onclick="toggleAll(false)">Collapse all</button>
  </div>

  <div class="container" id="container">
    __BODY__
  </div>

  <footer>
    __TEAM__
    <div class="hint">Photos: place files in <b>assets/photos/</b> named <b>xiaowen.wang.svg</b>, <b>yixuan.fan.svg</b>, <b>kay.sun.svg</b> (svg/png/jpg). Missing photos will show a placeholder.</div>
  </footer>

  <script>
    function toggleAll(open) {
      document.querySelectorAll('details.group, details.country').forEach(d => d.open = open);
    }

    function filterList() {
      const q = (document.getElementById('search').value || '').toLowerCase().trim();
      const groups = document.querySelectorAll('details.group');

      groups.forEach(g => {
        let anyVisible = false;
        const countries = g.querySelectorAll('details.country');
        countries.forEach(c => {
          const hay = (c.getAttribute('data-hay') || '').toLowerCase();
          const show = (!q) || hay.includes(q);
          c.style.display = show ? '' : 'none';
          if (show) anyVisible = true;
        });
        g.style.display = anyVisible ? '' : 'none';
        if (q && anyVisible) g.open = true;
      });
    }
  </script>
</body>
</html>
"""
    return (tpl
            .replace("__PAGE_TITLE__", escape(page_title))
            .replace("__H1__", escape(page_title))
            .replace("__SUBTITLE__", escape(subtitle))
            .replace("__NAV__", nav_html)
            .replace("__BODY__", body_html)
            .replace("__TEAM__", team_html)
            )


def build_team_html(placeholder_svg: str) -> str:
    parts = ["<div class=\"team\">", "<h3>Team</h3>", "<div class=\"members\">"]
    for m in TEAM:
        email_txt = escape(m["email"])

        # Second line: email and optional external links (e.g., LinkedIn)
        line2_parts = [f"<a href=\"mailto:{email_txt}\">{email_txt}</a>"]
        for l in m.get("links", []):
            line2_parts.append(
                f"<a href=\"{html.escape(l['href'], quote=True)}\" target=\"_blank\" rel=\"noopener\">{escape(l['label'])}</a>"
            )
        line2 = " <span class=\"sep\">·</span> ".join(line2_parts)

        photo = html.escape(m["photo"], quote=True)
        # Use onerror to fall back to placeholder svg (data URL)
        img = (
            f"<img src=\"{photo}\" alt=\"{escape(m['name'])}\" "
            f"onerror=\"this.onerror=null;this.src='{placeholder_svg}';\" />"
        )

        parts.append(
            "<div class=\"member\">"
            f"<div class=\"avatar\">{img}</div>"
            "<div class=\"mmeta\">"
            f"<div class=\"name\">{escape(m['name'])}</div>"
            f"<div class=\"roleline\">{escape(m['role'])}<br>{line2}</div>"
            "</div>"
            "</div>"
        )
    parts.append("</div></div>")
    return "".join(parts)


def build_nav_html(pages: List[Tuple[str, str]], active_href: str) -> str:
    # pages: [(label, href)]
    out = []
    for label, href in pages:
        cls = "active" if href == active_href else ""
        out.append(f"<a class=\"{cls}\" href=\"{html.escape(href, quote=True)}\">{escape(label)}</a>")
    return "".join(out)


def svg_to_inline(svg_text: str) -> str:
    # strip xml/doctype and ensure it can sit inline
    s = svg_text.strip()
    # remove xml declaration if present
    if s.startswith("<?xml"):
        s = s.split("?>", 1)[-1].lstrip()
    return s


def build_external_icon_svg() -> str:
    # simple inline external-link icon
    return (
        "<span class=\"ext\">"
        "<svg viewBox=\"0 0 24 24\" fill=\"none\" xmlns=\"http://www.w3.org/2000/svg\">"
        "<path d=\"M14 3h7v7\" stroke=\"currentColor\" stroke-width=\"2\" stroke-linecap=\"round\" stroke-linejoin=\"round\"/>"
        "<path d=\"M10 14L21 3\" stroke=\"currentColor\" stroke-width=\"2\" stroke-linecap=\"round\" stroke-linejoin=\"round\"/>"
        "<path d=\"M21 14v7H3V3h7\" stroke=\"currentColor\" stroke-width=\"2\" stroke-linecap=\"round\" stroke-linejoin=\"round\"/>"
        "</svg>"
        "</span>"
    )


def sheet_rows(ws, headers: List[str]) -> List[Dict[str, str]]:
    idx = {h: i for i, h in enumerate(headers)}

    def get(row, col):
        if col not in idx:
            return None
        v = row[idx[col]]
        return v

    rows_out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() for v in r):
            continue
        country = get(r, DEFAULT_COLUMNS["country"])
        if not country or not str(country).strip():
            continue
        rows_out.append({
            "country": str(country).strip(),
            "region": get(r, DEFAULT_COLUMNS["region"]),
            "food": get(r, DEFAULT_COLUMNS["food"]),
            "f_link": get(r, DEFAULT_COLUMNS["f_link"]),
            "supp": get(r, DEFAULT_COLUMNS["supp"]),
            "s_link": get(r, DEFAULT_COLUMNS["s_link"]),
            "drug": get(r, DEFAULT_COLUMNS["drug"]),
            "d_link": get(r, DEFAULT_COLUMNS["d_link"]),
            "feed": get(r, DEFAULT_COLUMNS["feed"]),
            "fe_link": get(r, DEFAULT_COLUMNS["fe_link"]),
        })
    return rows_out


def build_body(rows: List[Dict[str, str]], region_map: Dict[str, str], svg_yes: str, svg_no: str) -> str:
    ext_icon = build_external_icon_svg()

    # group
    grouped: Dict[str, List[Dict[str, str]]] = {k: [] for k in REGION_ORDER}
    for r in rows:
        region = classify_region(r["country"], r.get("region"), region_map)
        if region not in grouped:
            grouped["其他"].append(r)
        else:
            grouped[region].append(r)

    # sort countries within region for stability
    for reg in grouped:
        grouped[reg].sort(key=lambda x: x["country"])

    parts: List[str] = []
    for reg in REGION_ORDER:
        items = grouped.get(reg, [])
        if not items:
            continue
        parts.append(f"<details class=\"group\"><summary><span>{escape(reg)}</span><span class=\"badge\">{len(items)} countries</span></summary>")
        for r in items:
            food_html = render_status(r.get("food"), svg_yes, svg_no)
            supp_html = render_status(r.get("supp"), svg_yes, svg_no)
            drug_html = render_status(r.get("drug"), svg_yes, svg_no)
            feed_html = render_status(r.get("feed"), svg_yes, svg_no)

            f_link_html = render_link(r.get("f_link"), ext_icon)
            s_link_html = render_link(r.get("s_link"), ext_icon)
            d_link_html = render_link(r.get("d_link"), ext_icon)
            fe_link_html = render_link(r.get("fe_link"), ext_icon)

            # search haystack
            hay = " | ".join([
                r.get("country", "") or "",
                str(r.get("food") or ""),
                str(r.get("supp") or ""),
                str(r.get("drug") or ""),
                str(r.get("feed") or ""),
                str(r.get("f_link") or ""),
                str(r.get("s_link") or ""),
                str(r.get("d_link") or ""),
                str(r.get("fe_link") or ""),
            ])
            hay_attr = html.escape(hay, quote=True)

            parts.append(
                f"<details class=\"country\" data-hay=\"{hay_attr}\">"
                f"<summary><span>{escape(r['country'])}</span></summary>"
                "<div class=\"content\">"
                "<div class=\"grid\">"
                "<div class=\"cell-title\">食品</div>"
                "<div class=\"cell-title\">保健品/膳食补充剂</div>"
                "<div class=\"cell-title\">药品</div>"
                "<div class=\"cell-title\">动物食品</div>"
                f"<div class=\"cell\">{food_html}</div>"
                f"<div class=\"cell\">{supp_html}</div>"
                f"<div class=\"cell\">{drug_html}</div>"
                f"<div class=\"cell\">{feed_html}</div>"
                "</div>"
                "<div class=\"links\">"
                f"<div>{f_link_html}</div>"
                f"<div>{s_link_html}</div>"
                f"<div>{d_link_html}</div>"
                f"<div>{fe_link_html}</div>"
                "</div>"
                "</div>"
                "</details>"
            )
        parts.append("</details>")

    return "".join(parts)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", default="data.xlsx", help="Path to Excel file")
    ap.add_argument("--out", default=".", help="Output directory")
    args = ap.parse_args()

    out_dir = Path(args.out).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    here = Path(__file__).resolve().parent
    svg_yes = svg_to_inline(read_text(here / "assets/icons/yes.svg"))
    svg_no = svg_to_inline(read_text(here / "assets/icons/no.svg"))

    # photo placeholder as data URL svg
    placeholder_svg = (
        "data:image/svg+xml;utf8," +
        "<svg xmlns='http://www.w3.org/2000/svg' width='120' height='120'>"
        "<rect width='100%25' height='100%25' rx='18' ry='18' fill='%23f4f1ec'/>"
        "<circle cx='60' cy='48' r='18' fill='%23ddd6cc'/>"
        "<rect x='26' y='74' width='68' height='30' rx='15' fill='%23ddd6cc'/>"
        "</svg>"
    )

    region_map = load_region_map(here / "region_map.csv")

    wb = openpyxl.load_workbook(args.excel)
    sheets = wb.sheetnames

    # Define pages
    pages: List[Tuple[str, str]] = [("Home", "index.html")]
    for sn in sheets:
        href = f"{sn}.html"
        pages.append((sn, href))

    # Build each sheet page
    for sn in sheets:
        ws = wb[sn]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        # Ensure required headers exist
        missing = [DEFAULT_COLUMNS[k] for k in ("country","food","f_link","supp","s_link","drug","d_link","feed","fe_link") if DEFAULT_COLUMNS[k] not in headers]
        if missing:
            raise SystemExit(f"Sheet '{sn}' missing columns: {missing}")

        rows = sheet_rows(ws, headers)
        body = build_body(rows, region_map, svg_yes, svg_no)

        nav = build_nav_html(pages, f"{sn}.html")
        team_html = build_team_html(placeholder_svg)
        html_text = html_shell(page_title=sn, subtitle="WeCompliance", nav_html=nav, body_html=body, team_html=team_html)

        (out_dir / f"{sn}.html").write_text(html_text, encoding="utf-8")

    # Build index.html
    nav = build_nav_html(pages, "index.html")
    cards = [
        "<div class=\"container\">"
        "  <div class=\"sheetcards\">"
    ]
    for sn in sheets:
        icon_slug = slugify(sn)
        icon_path = f"assets/sheets/{icon_slug}.svg"
        cards.append(
            f"    <a class=\"sheetcard\" href=\"{html.escape(sn+'.html', quote=True)}\">"
            f"      <img class=\"img\" src=\"{html.escape(icon_path, quote=True)}\" alt=\"\" onerror=\"this.onerror=null;this.src='assets/sheets/default.svg';\">"
            "      <div class=\"meta\">"
            f"        <div class=\"t\">{escape(sn)}</div>"
            f"        <div class=\"d\">Open {escape(sn)} page</div>"
            "      </div>"
            "    </a>"
        )
    cards.append("  </div></div>")

    team_html = build_team_html(placeholder_svg)
    index_html = html_shell(page_title="WeCompliance", subtitle="Probiotics Compliance Advice Around the World", nav_html=nav, body_html="".join(cards), team_html=team_html)
    (out_dir / "index.html").write_text(index_html, encoding="utf-8")

    print("Generated:")
    print(" -", out_dir / "index.html")
    for sn in sheets:
        print(" -", out_dir / f"{sn}.html")


if __name__ == "__main__":
    main()
